"""
api/export.py — Excel and PDF export builders.

Provides:
  - ExcelBuilder: generates .xlsx with KPI sheets + Issues sheet
  - (future) PdfBuilder: generates executive PDF report
"""
from __future__ import annotations

import io
from datetime import date

import structlog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from storage.models import KPIResult, FactIssue

logger = structlog.get_logger(__name__)

HEADER_FILL = PatternFill(start_color="2e3248", end_color="2e3248", fill_type="solid")
HEADER_FONT = Font(bold=True, color="ffffff", size=11)
CATEGORY_FILLS = {
    "delivery": PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid"),
    "quality": PatternFill(start_color="3d1a1a", end_color="3d1a1a", fill_type="solid"),
    "risk": PatternFill(start_color="3d2e1a", end_color="3d2e1a", fill_type="solid"),
    "risk_control": PatternFill(start_color="3d2e1a", end_color="3d2e1a", fill_type="solid"),
    "data_quality": PatternFill(start_color="1a3d2e", end_color="1a3d2e", fill_type="solid"),
    "team": PatternFill(start_color="2e1a3d", end_color="2e1a3d", fill_type="solid"),
}

SHEET_CONFIG = [
    ("Delivery KPIs", "delivery"),
    ("Quality KPIs", "quality"),
    ("Risk KPIs", ("risk", "risk_control")),
    ("Data Quality KPIs", "data_quality"),
    ("Team KPIs", "team"),
]


class ExcelBuilder:
    """Builds an .xlsx workbook from KPI results and issues."""

    def __init__(self, project_key: str):
        self.project_key = project_key
        self.wb = Workbook()

    def add_kpi_sheet(self, sheet_name: str, kpis: list[KPIResult]) -> None:
        ws = self.wb.create_sheet(title=sheet_name)
        headers = ["KPI", "Period", "Value", "Previous", "Delta", "Trend", "Risk Level", "Interpretation"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for i, k in enumerate(kpis, start=2):
            ws.cell(row=i, column=1, value=k.kpi_name)
            ws.cell(row=i, column=2, value=k.period_label)
            val_cell = ws.cell(row=i, column=3)
            val_cell.value = round(k.current_value, 1) if k.current_value is not None else ""
            val_cell.number_format = "#,##0.0"
            prev_cell = ws.cell(row=i, column=4)
            prev_cell.value = round(k.previous_value, 1) if k.previous_value is not None else ""
            prev_cell.number_format = "#,##0.0"
            delta_cell = ws.cell(row=i, column=5)
            delta_cell.value = round(k.delta, 2) if k.delta is not None else ""
            delta_cell.number_format = "+#,##0.00;-#,##0.00"
            ws.cell(row=i, column=6, value=k.trend.value if k.trend else "")
            ws.cell(row=i, column=7, value=k.risk_level.value if k.risk_level else "")
            ws.cell(row=i, column=8, value=k.interpretation or "")

        self._auto_width(ws)

    def add_issues_sheet(self, issues: list[FactIssue]) -> None:
        ws = self.wb.create_sheet(title="Issues")
        headers = [
            "Key", "Summary", "Type", "Status", "Priority", "Assignee",
            "Created", "Resolved", "Age (d)", "Resolution (d)",
            "Overdue", "Reopened", "Missing Assignee", "Missing Priority",
        ]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for i, r in enumerate(issues, start=2):
            ws.cell(row=i, column=1, value=r.jira_key)
            ws.cell(row=i, column=2, value=r.summary or "")
            ws.cell(row=i, column=3, value=r.issue_type or "")
            ws.cell(row=i, column=4, value=r.status or "")
            ws.cell(row=i, column=5, value=r.priority or "")
            ws.cell(row=i, column=6, value=r.assignee_id or "")
            ws.cell(row=i, column=7, value=r.created_date.isoformat() if r.created_date else "")
            ws.cell(row=i, column=8, value=r.resolved_date.isoformat() if r.resolved_date else "")
            ws.cell(row=i, column=9, value=r.age_days or "")
            res_cell = ws.cell(row=i, column=10)
            res_cell.value = round(r.resolution_time_days, 1) if r.resolution_time_days else ""
            res_cell.number_format = "#,##0.0"
            ws.cell(row=i, column=11, value="Yes" if r.is_overdue else "No")
            ws.cell(row=i, column=12, value=r.times_reopened or 0)
            ws.cell(row=i, column=13, value="Yes" if r.dq_missing_assignee else "No")
            ws.cell(row=i, column=14, value="Yes" if r.dq_missing_priority else "No")

        self._auto_width(ws)

    @staticmethod
    def _auto_width(ws) -> None:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    def build(self) -> io.BytesIO:
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        buf = io.BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf


async def build_xlsx(
    project_key: str,
    kpis: list[KPIResult],
    issues: list[FactIssue],
) -> io.BytesIO:
    """Build an .xlsx workbook with KPI category sheets + Issues sheet."""
    builder = ExcelBuilder(project_key)

    # Group KPIs by category
    from collections import defaultdict
    by_cat: dict[str, list[KPIResult]] = defaultdict(list)
    for k in kpis:
        by_cat[k.kpi_category].append(k)

    for sheet_name, cat_key in SHEET_CONFIG:
        if isinstance(cat_key, tuple):
            sheet_kpis = []
            for ck in cat_key:
                sheet_kpis.extend(by_cat.get(ck, []))
        else:
            sheet_kpis = by_cat.get(cat_key, [])
        if sheet_kpis:
            builder.add_kpi_sheet(sheet_name, sheet_kpis)

    if issues:
        builder.add_issues_sheet(issues)

    return builder.build()


# ---------------------------------------------------------------------------
# PDF Export
# ---------------------------------------------------------------------------


def _risk_classify(score: float | None) -> str:
    if score is None:
        return "unknown"
    if score >= 75:
        return "critical"
    if score >= 50:
        return "high"
    if score >= 25:
        return "medium"
    return "low"


async def build_pdf(project_key: str) -> io.BytesIO:
    """Build an A4 PDF executive report."""
    import json
    from pathlib import Path
    from weasyprint import HTML
    from sqlalchemy import select, func, desc
    from storage.database import get_db
    from storage.models import DimProject, FactIssue, KPIResult, RiskScore

    template_path = Path(__file__).parent / "export_templates" / "report.html"
    template = template_path.read_text(encoding="utf-8")

    today = date.today()

    async with get_db() as db:
        proj = await db.get(DimProject, project_key)
        project_name = proj.name if proj else project_key

        # Latest risk score (1m period)
        risk_row = (await db.execute(
            select(RiskScore).where(
                RiskScore.project_key == project_key,
                RiskScore.period_label == "1m",
            ).order_by(RiskScore.calculation_date.desc()).limit(1)
        )).scalar_one_or_none()

        # Issue counts
        total_open = (await db.execute(
            select(func.count(FactIssue.id)).where(
                FactIssue.project_key == project_key,
                FactIssue.status_category != "Done",
            )
        )).scalar() or 0

        total_overdue = (await db.execute(
            select(func.count(FactIssue.id)).where(
                FactIssue.project_key == project_key,
                FactIssue.is_overdue == True,
            )
        )).scalar() or 0

        critical_open = (await db.execute(
            select(func.count(FactIssue.id)).where(
                FactIssue.project_key == project_key,
                FactIssue.status_category != "Done",
                FactIssue.priority.in_(["Critical", "Blocker", "Highest"]),
            )
        )).scalar() or 0

        unassigned_open = (await db.execute(
            select(func.count(FactIssue.id)).where(
                FactIssue.project_key == project_key,
                FactIssue.status_category != "Done",
                FactIssue.assignee_id == None,
            )
        )).scalar() or 0

        # Top KPIs (latest 1m period)
        kpi_rows_raw = (await db.execute(
            select(KPIResult).where(
                KPIResult.project_key == project_key,
                KPIResult.period_label == "1m",
            ).order_by(KPIResult.kpi_category, KPIResult.kpi_name)
        )).scalars().all()

        # Alerts from executive summary logic
        alerts = _build_pdf_alerts(total_overdue, critical_open, unassigned_open,
                                   risk_row.composite_risk if risk_row else 0)

    # Risk scores
    composite = round(risk_row.composite_risk, 1) if risk_row else 0
    delivery = round(risk_row.delivery_risk, 1) if risk_row else 0
    quality = round(risk_row.quality_risk, 1) if risk_row else 0
    compliance = round(risk_row.compliance_risk, 1) if risk_row else 0
    operational = round(risk_row.operational_risk, 1) if risk_row else 0

    # KPI table rows
    kpi_rows = ""
    for k in kpi_rows_raw[:15]:
        val = round(k.current_value, 1) if k.current_value is not None else "—"
        trend = k.trend.value if k.trend else "unknown"
        risk_lvl = k.risk_level.value if k.risk_level else "low"
        kpi_rows += (
            f"<tr><td>{k.kpi_name}</td><td>{val}</td>"
            f"<td class='trend-{trend}'>{trend}</td>"
            f"<td class='risk-{risk_lvl}'>{risk_lvl}</td></tr>\n"
        )

    # Alert items
    alert_items = ""
    for a in alerts[:8]:
        level = a.get("level", "medium")
        msg = a.get("message", "")
        alert_items += f"<li class='alert-{level}'><strong>{level.upper()}:</strong> {msg}</li>\n"
    if not alert_items:
        alert_items = "<li style='color:#666'>No active alerts. Project health is good.</li>"

    html = template.format(
        project_name=project_name,
        generated_at=today.isoformat(),
        overall_risk_level=_risk_classify(composite),
        composite_score=composite,
        composite_risk_level=_risk_classify(composite),
        delivery_score=delivery,
        delivery_level=_risk_classify(delivery),
        quality_score=quality,
        quality_level=_risk_classify(quality),
        compliance_score=compliance,
        compliance_level=_risk_classify(compliance),
        operational_score=operational,
        operational_level=_risk_classify(operational),
        total_open=total_open,
        total_overdue=total_overdue,
        critical_open=critical_open,
        unassigned_open=unassigned_open,
        kpi_rows=kpi_rows,
        alert_items=alert_items,
    )

    buf = io.BytesIO()
    HTML(string=html).write_pdf(buf)
    buf.seek(0)
    return buf


def _build_pdf_alerts(overdue: int, critical: int, unassigned: int, risk: float) -> list[dict]:
    alerts = []
    if critical >= 10:
        alerts.append({"level": "critical", "message": f"{critical} critical issues open — immediate attention"})
    elif critical >= 3:
        alerts.append({"level": "high", "message": f"{critical} critical/blocker issues open"})
    if overdue >= 20:
        alerts.append({"level": "high", "message": f"{overdue} overdue issues across all projects"})
    elif overdue >= 5:
        alerts.append({"level": "medium", "message": f"{overdue} overdue issues need re-planning"})
    if unassigned >= 20:
        alerts.append({"level": "high", "message": f"{unassigned} open issues have no assignee"})
    if risk >= 75:
        alerts.append({"level": "critical", "message": "Risk is CRITICAL — escalate to leadership"})
    elif risk >= 50:
        alerts.append({"level": "high", "message": "Risk is HIGH — review in next steering committee"})
    return alerts
